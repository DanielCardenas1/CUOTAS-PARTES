#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generador de formato Excel exacto basado en HOS_890201488.xlsx
Replica el formato oficial de liquidación de cuotas partes
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from app.db import get_session
from sqlalchemy import text

def obtener_dtf_mes(año, mes):
    """Obtiene la DTF para un mes específico"""
    session = get_session()
    
    try:
        fecha_consulta = date(año, mes, 1)
        query = text("SELECT tasa FROM dtf_mensual WHERE periodo = :fecha")
        result = session.execute(query, {'fecha': fecha_consulta}).fetchone()
        
        if result:
            dtf_decimal = float(result[0])
            dtf_porcentaje = dtf_decimal * 100
            return dtf_porcentaje
        else:
            return 10.0
            
    except Exception as e:
        print(f"Error consultando DTF: {e}")
        return 10.0
    finally:
        session.close()

def obtener_ipc_desde_bd(año_inicio, año_fin):
    """Obtiene el IPC acumulado desde BD"""
    session = get_session()
    
    try:
        query = text("SELECT anio, valor FROM ipc_anual WHERE anio >= :inicio AND anio <= :fin ORDER BY anio")
        result = session.execute(query, {'inicio': año_inicio, 'fin': año_fin}).fetchall()
        
        ipc_acumulado = 1.0
        for row in result:
            ipc_valor = float(row[1])
            ipc_acumulado *= ipc_valor
        
        return ipc_acumulado
    
    except Exception as e:
        print(f"Error consultando IPC: {e}")
        return 1.0
    
    finally:
        session.close()

def ajustar_base_por_ipc(base_2025, año_cuenta):
    """Ajusta la base de cálculo de 2025 al año de la cuenta usando datos reales de BD"""
    if año_cuenta >= 2025:
        return base_2025
    
    ipc_acumulado = obtener_ipc_desde_bd(año_cuenta, 2025)
    base_ajustada = base_2025 / ipc_acumulado
    
    return base_ajustada

def calcular_interes_mensual_formato_oficial(capital_fijo, fecha_cuenta, fecha_corte):
    """Calcula interés usando DTF y días del MISMO mes de la cuenta"""
    if fecha_cuenta >= fecha_corte:
        return 0.0
    
    # DTF del mismo mes de la cuenta
    dtf_ea = obtener_dtf_mes(fecha_cuenta.year, fecha_cuenta.month) / 100
    
    # Días del mismo mes de la cuenta
    if fecha_cuenta.month == 12:
        fecha_fin_mes = date(fecha_cuenta.year + 1, 1, 1) - relativedelta(days=1)
    else:
        fecha_fin_mes = date(fecha_cuenta.year, fecha_cuenta.month + 1, 1) - relativedelta(days=1)
    
    dias_mes = fecha_fin_mes.day
    
    # Fórmula Excel: Capital × ((1+DTF)^(días/365) - 1)
    if dias_mes > 0:
        factor_interes = ((1 + dtf_ea) ** (dias_mes / 365)) - 1
        interes_mes = capital_fijo * factor_interes
        return round(interes_mes, 2)
    
    return 0.0

def generar_datos_liquidacion():
    """Genera los datos de liquidación en el formato requerido"""
    
    # Obtener pensionado
    session = get_session()
    try:
        query = text('''SELECT identificacion, nombre, numero_mesadas, 
                        fecha_ingreso_nomina, empresa, base_calculo_cuota_parte,
                        porcentaje_cuota_parte
                        FROM pensionado LIMIT 1''')
        result = session.execute(query).fetchone()
        pensionado = result
    finally:
        session.close()
    
    fecha_corte = date(2025, 8, 31)
    base_calculo = float(pensionado[5])
    porcentaje_cuota_parte = float(pensionado[6]) if pensionado[6] else 0.7611
    
    # Generar datos por mes (usar la misma lógica que el sistema funcionando)
    datos_meses = []
    fecha_inicial = date(2022, 9, 1)
    
    for i in range(36):
        fecha_mes = fecha_inicial + relativedelta(months=i)
        año = fecha_mes.year
        mes = fecha_mes.month
        
        # Base ajustada por IPC (igual que el sistema funcionando)
        base_ajustada_año = ajustar_base_por_ipc(base_calculo, año)
        capital_fijo = base_ajustada_año * porcentaje_cuota_parte
        
        # Prima si aplica
        prima = 0
        capital_final = capital_fijo
        mesadas = int(pensionado[2])
        if (mesadas == 13 and mes == 12) or (mesadas == 14 and mes in [6, 12]):
            prima = capital_fijo
            capital_final = capital_fijo + prima
        
        # Interés (usar la misma función que el sistema funcionando)
        interes = calcular_interes_mensual_formato_oficial(capital_final, fecha_mes, fecha_corte)
        
        # DTF y días para mostrar
        dtf_porcentaje = obtener_dtf_mes(año, mes)
        if mes == 12:
            fecha_fin_mes = date(año + 1, 1, 1) - relativedelta(days=1)
        else:
            fecha_fin_mes = date(año, mes + 1, 1) - relativedelta(days=1)
        dias = fecha_fin_mes.day
        
        datos_meses.append({
            'año': año,
            'mes': mes,
            'fecha': fecha_mes,
            'capital_fijo': capital_fijo,
            'prima': prima,
            'capital_total': capital_final,
            'dias': dias,
            'dtf_ea': dtf_porcentaje,
            'interes': interes,
            'total_cuenta': capital_final + interes
        })
    
    return pensionado, datos_meses

def crear_formato_excel_oficial():
    """Crea el archivo Excel con formato oficial exacto"""
    
    pensionado, datos_meses = generar_datos_liquidacion()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LIQUIDACION CUOTAS PARTES"
    
    # Estilos
    font_title = Font(name='Arial', size=12, bold=True)
    font_header = Font(name='Arial', size=10, bold=True)
    font_normal = Font(name='Arial', size=9)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_total = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # ENCABEZADO PRINCIPAL
    ws.merge_cells('A1:I1')
    ws['A1'] = f"{pensionado[4]}"  # Empresa
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    
    ws.merge_cells('A2:I2')
    ws['A2'] = "LIQUIDACIÓN POR CUOTAS PARTES ART. 4 LEY 100 DE 1993"
    ws['A2'].font = font_header
    ws['A2'].alignment = align_center
    
    # INFORMACIÓN DEL PENSIONADO
    ws['A4'] = "Nombre:"
    ws['B4'] = pensionado[1]
    ws['G4'] = "Resolución:"
    ws['H4'] = "1756 de 2006"
    
    ws['A5'] = "Identificación:"
    ws['B5'] = pensionado[0]
    ws['G5'] = "Fecha resolución:"
    
    ws['A6'] = "Empresa:"
    ws['G6'] = "Resolución:"
    
    ws['A7'] = "Identificación:"
    ws['G7'] = ""
    
    # TABLA RESUMEN POR AÑO (Como en el Excel original)
    row = 10
    ws[f'A{row}'] = "Ingreso a Nómina"
    ws[f'B{row}'] = "% Cuota Parte"
    ws[f'C{row}'] = "Salud"
    ws[f'D{row}'] = "Cuota parte actual"
    ws[f'E{row}'] = "Capital pendiente"
    ws[f'F{row}'] = "Intereses Devengados"
    ws[f'G{row}'] = "Estado"
    ws[f'H{row}'] = "Pagos a Cuenta"
    
    for col in range(1, 9):
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = align_center
    
    # TABLA DETALLADA MENSUAL - CABECERAS
    row = 20
    headers = ["Año", "Al Servicio", "Al Hasta", "Prima", "Pensión por", "Pensión", 
               "Base Cálculo anual", "Vr. Cuota parte", "Periodos", "Total Cuenta", "Pagos", "Saldo Pendiente"]
    
    for col, header in enumerate(headers, 1):
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = align_center
    
    # TABLA DETALLADA CON INTERESES DTF - CABECERAS  
    row = 30
    ws[f'A{row}'] = "PERIODO"
    ws[f'B{row}'] = "CON INTERESES DTF"
    ws.merge_cells(f'B{row}:I{row}')
    
    row = 31
    headers_dtf = ["", "Periodos", "Días", "DTF EA", "Valor Intereses", "Int.x Cuota Acumulada"]
    
    for col, header in enumerate(headers_dtf, 1):
        if header:  # Skip empty first column
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.value = header
            cell.font = font_header
            cell.fill = fill_header
            cell.border = border_thin
            cell.alignment = align_center
    
    # DATOS MENSUALES
    meses_nombres = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                     "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    
    row = 32
    total_intereses = 0
    
    for dato in datos_meses:
        mes_nombre = meses_nombres[dato['mes'] - 1]
        año = dato['año']
        
        # Período
        ws[f'A{row}'] = f"{mes_nombre}-{año}"
        ws[f'A{row}'].font = font_normal
        ws[f'A{row}'].alignment = align_left
        
        # Días
        ws[f'C{row}'] = dato['dias']
        ws[f'C{row}'].font = font_normal
        ws[f'C{row}'].alignment = align_center
        
        # DTF EA
        ws[f'D{row}'] = f"{dato['dtf_ea']:.2f}%"
        ws[f'D{row}'].font = font_normal
        ws[f'D{row}'].alignment = align_center
        
        # Valor Intereses
        ws[f'E{row}'] = dato['interes']
        ws[f'E{row}'].font = font_normal
        ws[f'E{row}'].alignment = align_right
        ws[f'E{row}'].number_format = '"$"#,##0.00'
        
        # Total acumulado
        total_intereses += dato['interes']
        ws[f'F{row}'] = total_intereses
        ws[f'F{row}'].font = font_normal
        ws[f'F{row}'].alignment = align_right
        ws[f'F{row}'].number_format = '"$"#,##0.00'
        
        # Bordes para toda la fila
        for col in range(1, 7):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.border = border_thin
        
        row += 1
    
    # FILA DE TOTALES
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = font_header
    ws[f'A{row}'].fill = fill_total
    ws[f'A{row}'].alignment = align_center
    
    ws[f'E{row}'] = total_intereses
    ws[f'E{row}'].font = font_header
    ws[f'E{row}'].fill = fill_total
    ws[f'E{row}'].alignment = align_right
    ws[f'E{row}'].number_format = '"$"#,##0.00'
    
    for col in range(1, 7):
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.border = border_thin
        if col in [1, 5]:  # Solo A y E tienen contenido
            cell.fill = fill_total
    
    # NOTAS AL FINAL
    row += 3
    ws[f'A{row}'] = f"Total capital correspondiente a los 36 cuentas desde 01 de sep de 2022 a 31 de jul de 2025"
    ws[f'A{row}'].font = font_normal
    
    row += 1
    ws[f'A{row}'] = f"Intereses causados (Ley 100 de 1993) de agosto 2022 a julio de 2025"
    ws[f'A{row}'].font = font_normal
    
    row += 1
    total_final = sum(d['capital_total'] for d in datos_meses) + total_intereses
    ws[f'A{row}'] = f"Total de la cuenta con intereses julio de 2025"
    ws[f'B{row}'] = total_final
    ws[f'B{row}'].number_format = '"$"#,##0.00'
    ws[f'B{row}'].font = Font(name='Arial', size=9, bold=True)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    
    # Guardar archivo
    nombre_archivo = f"LIQUIDACION_FORMATO_OFICIAL_{pensionado[0]}_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(nombre_archivo)
    
    print(f"✅ Archivo creado: {nombre_archivo}")
    print(f"📊 Total intereses: ${total_intereses:,.2f}")
    print(f"💰 Total liquidación: ${total_final:,.2f}")
    
    return nombre_archivo

if __name__ == "__main__":
    crear_formato_excel_oficial()
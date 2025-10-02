#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generador de formato Excel oficial - VERSIÓN CORREGIDA
Usa directamente las funciones del sistema funcionando
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Importar funciones del sistema funcionando
from mostrar_liquidacion_36 import (
    obtener_dtf_mes,
    ajustar_base_por_ipc,
    calcular_interes_mensual_unico,
    generar_36_cuentas_pensionado
)

from app.db import get_session
from sqlalchemy import text

def crear_excel_formato_oficial():
    """Crea el Excel con formato oficial usando el sistema funcionando"""
    
    # Obtener pensionado (mismo que el sistema funcionando)
    session = get_session()
    try:
        query = text('''SELECT identificacion, nombre, numero_mesadas, 
                        fecha_ingreso_nomina, empresa, base_calculo_cuota_parte,
                        porcentaje_cuota_parte 
                        FROM pensionado LIMIT 1''')
        result = session.execute(query).fetchone()
        pensionado = result
    finally:
        session.close()
    
    # Usar la función que ya funciona correctamente
    fecha_corte = date(2025, 8, 31)
    cuentas = generar_36_cuentas_pensionado(pensionado, fecha_corte)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LIQUIDACION CUOTAS PARTES"
    
    # Estilos
    font_title = Font(name='Arial', size=12, bold=True)
    font_header = Font(name='Arial', size=10, bold=True)
    font_normal = Font(name='Arial', size=9)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_total = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # ENCABEZADO PRINCIPAL
    ws.merge_cells('A1:I1')
    ws['A1'] = f"{pensionado[4]}"  # Empresa
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    
    ws.merge_cells('A2:I2')
    ws['A2'] = "LIQUIDACIÓN POR CUOTAS PARTES ART. 4 LEY 100 DE 1993"
    ws['A2'].font = font_header
    ws['A2'].alignment = align_center
    
    # INFORMACIÓN DEL PENSIONADO
    ws['A4'] = "Nombre:"
    ws['B4'] = pensionado[1]
    ws['G4'] = "Resolución:"
    ws['H4'] = "1756 de 2006"
    
    ws['A5'] = "Identificación:"
    ws['B5'] = pensionado[0]
    ws['G5'] = "Fecha resolución:"
    
    # TABLA DETALLADA CON INTERESES DTF - CABECERAS  
    row = 10
    ws[f'A{row}'] = "PERIODO"
    ws[f'B{row}'] = "CON INTERESES DTF"
    ws.merge_cells(f'B{row}:F{row}')
    ws[f'B{row}'].font = font_header
    ws[f'B{row}'].fill = fill_header
    ws[f'B{row}'].alignment = align_center
    ws[f'B{row}'].border = border_thin
    
    row = 11
    headers_dtf = ["PERIODO", "Días", "DTF EA", "Valor Intereses", "Int.x Cuota Acumulada", "Capital"]
    
    for col, header in enumerate(headers_dtf, 1):
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = align_center
    
    # DATOS MENSUALES
    meses_nombres = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                     "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    
    row = 12
    total_intereses = 0
    total_capital = 0
    
    for cuenta in cuentas:
        mes_nombre = meses_nombres[cuenta['mes'] - 1]
        año = cuenta['año']
        
        # Período
        ws[f'A{row}'] = f"{mes_nombre}-{año}"
        ws[f'A{row}'].font = font_normal
        ws[f'A{row}'].alignment = align_left
        ws[f'A{row}'].border = border_thin
        
        # Días
        ws[f'B{row}'] = cuenta['dias_interes']
        ws[f'B{row}'].font = font_normal
        ws[f'B{row}'].alignment = align_center
        ws[f'B{row}'].border = border_thin
        
        # DTF EA
        ws[f'C{row}'] = f"{cuenta['dtf_interes']:.2f}%"
        ws[f'C{row}'].font = font_normal
        ws[f'C{row}'].alignment = align_center
        ws[f'C{row}'].border = border_thin
        
        # Valor Intereses
        ws[f'D{row}'] = cuenta['interes']
        ws[f'D{row}'].font = font_normal
        ws[f'D{row}'].alignment = align_right
        ws[f'D{row}'].number_format = '"$"#,##0.00'
        ws[f'D{row}'].border = border_thin
        
        # Total acumulado intereses
        total_intereses += cuenta['interes']
        ws[f'E{row}'] = total_intereses
        ws[f'E{row}'].font = font_normal
        ws[f'E{row}'].alignment = align_right
        ws[f'E{row}'].number_format = '"$"#,##0.00'
        ws[f'E{row}'].border = border_thin
        
        # Capital
        ws[f'F{row}'] = cuenta['capital']
        ws[f'F{row}'].font = font_normal
        ws[f'F{row}'].alignment = align_right
        ws[f'F{row}'].number_format = '"$"#,##0.00'
        ws[f'F{row}'].border = border_thin
        
        total_capital += cuenta['capital']
        row += 1
    
    # FILA DE TOTALES
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = font_header
    ws[f'A{row}'].fill = fill_total
    ws[f'A{row}'].alignment = align_center
    ws[f'A{row}'].border = border_thin
    
    # Total intereses
    ws[f'D{row}'] = total_intereses
    ws[f'D{row}'].font = font_header
    ws[f'D{row}'].fill = fill_total
    ws[f'D{row}'].alignment = align_right
    ws[f'D{row}'].number_format = '"$"#,##0.00'
    ws[f'D{row}'].border = border_thin
    
    # Total capital
    ws[f'F{row}'] = total_capital
    ws[f'F{row}'].font = font_header
    ws[f'F{row}'].fill = fill_total
    ws[f'F{row}'].alignment = align_right
    ws[f'F{row}'].number_format = '"$"#,##0.00'
    ws[f'F{row}'].border = border_thin
    
    # Completar bordes de totales
    for col in range(1, 7):
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.border = border_thin
        if col in [1, 4, 6]:  # Solo celdas con contenido
            cell.fill = fill_total
    
    # NOTAS AL FINAL
    row += 3
    ws[f'A{row}'] = f"Total capital correspondiente a las 30 cuentas de los últimos 30 meses"
    ws[f'A{row}'].font = font_normal
    ws[f'B{row}'] = total_capital
    ws[f'B{row}'].number_format = '"$"#,##0.00'
    ws[f'B{row}'].font = font_normal
    
    row += 1
    ws[f'A{row}'] = f"Intereses causados (Ley 100 de 1993) de sep 2022 a ago 2025"
    ws[f'A{row}'].font = font_normal
    ws[f'B{row}'] = total_intereses
    ws[f'B{row}'].number_format = '"$"#,##0.00'
    ws[f'B{row}'].font = font_normal
    
    row += 1
    total_final = total_capital + total_intereses
    ws[f'A{row}'] = f"Total de la cuenta con intereses agosto de 2025"
    ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True)
    ws[f'B{row}'] = total_final
    ws[f'B{row}'].number_format = '"$"#,##0.00'
    ws[f'B{row}'].font = Font(name='Arial', size=9, bold=True)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    
    # Guardar archivo
    nombre_archivo = f"LIQUIDACION_FORMATO_OFICIAL_{pensionado[0]}_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(nombre_archivo)
    
    print(f"✅ Archivo Excel creado: {nombre_archivo}")
    print(f"📊 Total capital: ${total_capital:,.2f}")
    print(f"💰 Total intereses: ${total_intereses:,.2f}")
    print(f"🎯 Total liquidación: ${total_final:,.2f}")
    print(f"\n📋 Resumen:")
    print(f"   - Pensionado: {pensionado[1]}")
    print(f"   - Identificación: {pensionado[0]}")
    print(f"   - Período: Últimos 30 meses (mes vencido)")
    print(f"   - Formato: Exacto como HOS_890201488.xlsx")
    
    return nombre_archivo

if __name__ == "__main__":
    crear_excel_formato_oficial()
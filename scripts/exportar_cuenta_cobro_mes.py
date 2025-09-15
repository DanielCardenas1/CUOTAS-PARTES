
import os
import argparse
from datetime import date, datetime
import pandas as pd
from sqlalchemy import create_engine
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

# Configuración
DB_USER = 'root'
DB_PASS = 'BeMaster.1'
DB_HOST = 'localhost'
DB_PORT = '3306'
DB_NAME = 'liquidaciones'
OUTPUT_DIR = 'reportes_liquidacion'
os.makedirs(OUTPUT_DIR, exist_ok=True)


# Columnas oficiales y orden (sin OBSERVACIONES)
COLUMNAS = [
    "No. Cédula",
    "Apellidos y Nombres",
    "SUSTITUTO",
    "No. DOCUMENTO SUSTITUTO",
    "% DE CONCURRENCIA",
    "VALOR MESADA",
    "PERIODO LIQUIDADO (INICIO)",
    "PERIODO LIQUIDADO (FIN)",
    "CAPITAL",
    "INTERESES",
    "TOTAL",
]



# Crear engine antes de usarlo
engine = create_engine(f"mysql+mysqlconnector://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

# --- Argumentos CLI ---
parser = argparse.ArgumentParser(description="Exportar cuenta de cobro mensual a PDF/Excel")
parser.add_argument('--pensionado', type=str, help='Cédula del pensionado a exportar (opcional)')
parser.add_argument('--solo_mes', type=str, help='YYYY-MM-01: solo ese mes (opcional)')
args = parser.parse_args()

df = pd.read_sql("SELECT * FROM v_cuenta_cobro_mes", engine)
df = df[COLUMNAS].copy()

# Eliminar registros de prueba antes de filtrar
df = df[~df["Apellidos y Nombres"].str.contains("prueba", case=False, na=False)].copy()

# Filtrar por cédula si se pasa --pensionado
if args.pensionado:
    df = df[df["No. Cédula"].astype(str) == str(args.pensionado)].copy()
    if df.empty:
        print(f"No se encontró pensionado con cédula {args.pensionado}.")
        exit(1)

# Filtrar por mes si se pasa --solo_mes (PERIODO LIQUIDADO (INICIO) == mes)
if args.solo_mes:
    try:
        mes_dt = pd.to_datetime(args.solo_mes)
        df = df[pd.to_datetime(df["PERIODO LIQUIDADO (INICIO)"]).dt.to_period('M') == mes_dt.to_period('M')].copy()
        if df.empty:
            print(f"No hay datos para el mes {args.solo_mes}.")
            exit(1)
    except Exception as e:
        print(f"Error en el formato de --solo_mes: {e}")
        exit(1)

if df.empty:
    print("No hay datos para exportar tras aplicar los filtros.")
    exit(1)


# No es necesario renombrar columnas si la vista ya las entrega correctamente


# Ya se eliminaron los registros de prueba y se aplicaron los filtros opcionales.


# --- Agrupar por entidad y generar carpetas ---
if 'nit_entidad' in df.columns:
    entidad_col = 'nit_entidad'
elif 'NIT' in df.columns:
    entidad_col = 'NIT'
else:
    raise Exception('No se encontró columna de NIT de entidad en la vista.')

for nit, df_ent in df.groupby(entidad_col):
    # Crear carpeta por entidad
    entidad_dir = os.path.join(OUTPUT_DIR, str(nit))
    os.makedirs(entidad_dir, exist_ok=True)
    # Por cada pensionado de la entidad
    for idx, (cedula, df_pens) in enumerate(df_ent.groupby("No. Cédula")):
        nombre_pens = df_pens["Apellidos y Nombres"].iloc[0]
        # Excel
        excel_out = os.path.join(entidad_dir, f"LIQUIDACION_MENSUAL_{cedula}_{date.today().isoformat()}.xlsx")
        with pd.ExcelWriter(excel_out, engine="openpyxl") as writer:
            df_pens.to_excel(writer, index=False, sheet_name="LIQUIDACION")
            ws = writer.sheets["LIQUIDACION"]
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            thin = Side(border_style="thin", color="000000")
            grid = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_font = Font(name="Arial", size=6.5, bold=True)
            cell_font = Font(name="Arial", size=6.5)
            for c in range(1, ws.max_column+1):
                cell = ws.cell(row=1, column=c)
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = grid
            for r in range(2, ws.max_row+1):
                ws.row_dimensions[r].height = 14.5
                for c in range(1, ws.max_column+1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = cell_font
                    cell.border = grid
            widths = [18, 28, 16, 18, 10, 13, 14, 12, 12, 14, 13, 10]
            for idx2, w in enumerate(widths, start=1):
                ws.column_dimensions[chr(64+idx2)].width = w
            for r in range(2, ws.max_row+1):
                for c in range(1, ws.max_column+1):
                    cell = ws.cell(row=r, column=c)
                    header = ws.cell(row=1, column=c).value
                    if header in ("CAPITAL", "INTERESES", "TOTAL", "VALOR MESADA"):
                        try:
                            cell.value = float(cell.value) if cell.value not in ("", None, "TOTAL") else cell.value
                        except Exception:
                            pass
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif header == "% DE CONCURRENCIA":
                        try:
                            val = float(cell.value)
                            cell.value = val / 100.0
                        except Exception:
                            pass
                        cell.number_format = '0.0000%'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif header in ("PERIODO LIQUIDADO (INICIO)", "PERIODO LIQUIDADO (FIN)"):
                        cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif header in ("No. Cédula", "No. DOCUMENTO SUSTITUTO"):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif header in ("Apellidos y Nombres", "SUSTITUTO", "OBSERVACIONES"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            total_row = ws.max_row
            ws.cell(row=total_row, column=2).font = Font(name="Arial", size=9.5, bold=True)
            for col in (9, 10, 11):
                ws.cell(row=total_row, column=col).font = Font(name="Arial", size=9.5, bold=True)
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_margins.left = 0.3
            ws.page_margins.right = 0.3
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            ws.freeze_panes = "A2"
        print(f"Excel generado: {excel_out}")

        # PDF
        from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.pdfbase.pdfmetrics import stringWidth
        pdf_out = os.path.join(entidad_dir, f"LIQUIDACION_MENSUAL_{cedula}_{date.today().isoformat()}.pdf")
        PAGE = landscape(letter)
        LM = RM = 85
        TM = 36; BM = 36
        AREA_W = PAGE[0] - (LM + RM)
        H = ParagraphStyle(
            "H", fontName="Helvetica-Bold", fontSize=7, leading=12,
            alignment=TA_CENTER
        )
        B = ParagraphStyle("B", fontName="Helvetica",      fontSize=6.5, leading=7.5)
        VAL = ParagraphStyle("VAL", fontName="Helvetica-Bold", fontSize=6.5, alignment=TA_CENTER)
        dfv = df_pens.copy()
        dfv.insert(0, "No.", range(1, len(dfv)+1))
        def _fmt_date(x):
            try:
                return pd.to_datetime(x).strftime("%d/%m/%Y")
            except Exception:
                return "" if x is None else str(x)
        dfv["PERIODO LIQUIDADO"] = (
            dfv["PERIODO LIQUIDADO (INICIO)"].apply(_fmt_date)
            + " – " +
            dfv["PERIODO LIQUIDADO (FIN)"].apply(_fmt_date)
        )
        COLS = [
            "No.",
            "Apellidos y Nombres",
            "No. Cédula",
            "SUSTITUTO",
            "No. DOCUMENTO SUSTITUTO",
            "% DE CONCURRENCIA",
            "VALOR MESADA",
            "PERIODO LIQUIDADO",
            "CAPITAL",
            "INTERESES",
            "TOTAL",
        ]
        dfv = dfv[COLS].copy()
        def _fmt_money(x):
            try:
                return f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except Exception:
                return "" if x is None else str(x)
        def _fmt_pct(x):
            try:
                return f"{float(x):.4f}%".replace(".", ",")
            except Exception:
                return ""
        for col in ["VALOR MESADA", "CAPITAL", "INTERESES", "TOTAL"]:
            dfv[col] = dfv[col].apply(_fmt_money)
        dfv["% DE CONCURRENCIA"] = dfv["% DE CONCURRENCIA"].apply(_fmt_pct)
        headers_text = [
            "No.",
            "APELLIDOS Y NOMBRES<br/>DEL PENSIONADO",
            "No.<br/>DOCUMENTO",
            "SUSTITUTO",
            "No.<br/>DOCUMENTO",
            "% DE<br/>CUOTA<br/>PARTE",
            "VALOR<br/>MESADA",
            "PERIODO LIQUIDADO",
            "CAPITAL",
            "INTERESES",
            "TOTAL",
        ]
        H = ParagraphStyle(
            "H",
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=14,
            alignment=TA_CENTER,
            spaceBefore=2,
            spaceAfter=2,
            splitLongWords=False,
            wordWrap="LTR",
        )
        headers = [Paragraph(t, H) for t in headers_text]
        pad_lr = 10
        def header_width_points(idx: int) -> float:
            lines = headers_text[idx].replace("<br/>", "\n").split("\n")
            return max(stringWidth(line, "Helvetica-Bold", 7) for line in lines) + pad_lr
        col_widths_pt = [
            28,   # No.
            85,  # Apellidos y Nombres
            62,   # No. Documento
            52,   # Sustituto
            62,   # No. Documento Sustituto
            50,   # % de Concurrencia
            54,   # Valor Mesada
            95,  # Periodo Liquidado
            65,   # Capital
            65,   # Intereses
            65,   # Total
        ]
        # Fila TOTAL fusionada
        total_capital   = float(pd.to_numeric(df_pens["CAPITAL"], errors="coerce").fillna(0).sum())
        total_intereses = float(pd.to_numeric(df_pens["INTERESES"], errors="coerce").fillna(0).sum())
        total_total     = float(pd.to_numeric(df_pens["TOTAL"], errors="coerce").fillna(0).sum())
        total_row = ["" ] * 11
        total_row[0] = "TOTAL"
        total_row[8] = _fmt_money(total_capital)
        total_row[9] = _fmt_money(total_intereses)
        total_row[10] = _fmt_money(total_total)
        total_row = [str(x) if not isinstance(x, Paragraph) else x for x in total_row]
        data = [headers] + dfv.fillna("").astype(str).values.tolist() + [total_row]
        tbl = Table(data, colWidths=col_widths_pt, repeatRows=1, hAlign="CENTER")
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 8),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("VALIGN", (0,0), (-1,0), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,0), 8),
            ("BOTTOMPADDING", (0,0), (-1,0), 8),
            ("FONTNAME",   (0,1), (-1,-2), "Helvetica"),
            ("FONTSIZE",   (0,1), (-1,-2), 7),
            ("VALIGN",     (0,1), (-1,-2), "MIDDLE"),
            ("TOPPADDING", (0,1), (-1,-2), 2),
            ("BOTTOMPADDING", (0,1), (-1,-2), 2),
            ("ALIGN", (0,1), (0,-2), "CENTER"),
            ("ALIGN", (1,1), (1,-2), "LEFT"),
            ("ALIGN", (2,1), (2,-2), "CENTER"),
            ("ALIGN", (3,1), (3,-2), "LEFT"),
            ("ALIGN", (4,1), (4,-2), "CENTER"),
            ("ALIGN", (5,1), (5,-2), "CENTER"),
            ("ALIGN", (6,1), (6,-2), "RIGHT"),
            ("ALIGN", (7,1), (7,-2), "CENTER"),
            ("ALIGN", (8,1), (10,-2), "RIGHT"),
            ("SPAN", (0,-1), (7,-1)),
            ("FONTNAME", (0,-1), (10,-1), "Helvetica-Bold"),
            ("FONTSIZE",  (0,-1), (10,-1), 8),
            ("ALIGN",     (0,-1), (0,-1),  "CENTER"),
            ("ALIGN",     (8,-1), (10,-1), "RIGHT"),
            ("TOPPADDING",(0,-1), (10,-1), 4),
            ("BOTTOMPADDING",(0,-1),(10,-1), 4),
            ("GRID", (0,0), (-1,-1), 0.6, colors.black),
            ("BOX",  (0,0), (-1,-1), 1.1, colors.black),
            ("LINEABOVE", (0,0), (-1,0), 1.1, colors.black),
            ("LINEBELOW", (0,0), (-1,0), 1.1, colors.black),
        ]))
        doc = BaseDocTemplate(
            pdf_out, pagesize=PAGE,
            leftMargin=LM, rightMargin=RM, topMargin=TM, bottomMargin=BM,
        )
        frame = Frame(LM, BM, AREA_W, PAGE[1] - (TM+BM), showBoundary=0)
        doc.addPageTemplates(PageTemplate(id="main", frames=[frame]))
        story = []
        story.append(Paragraph("LIQUIDACIÓN OFICIAL DE PENSIONADOS", ParagraphStyle("T", fontName="Helvetica-Bold", fontSize=12, alignment=TA_CENTER, spaceAfter=6)))
        from reportlab.platypus import KeepTogether
        story.append(KeepTogether(tbl))
        story.append(Spacer(1, 12))
        from num2words import num2words
        def numero_a_letras(valor):
            try:
                return num2words(valor, lang='es').replace('coma', 'punto').upper()
            except Exception:
                return str(valor)
        total_columna = pd.to_numeric(df_pens["TOTAL"], errors="coerce").fillna(0).sum()
        valor_letras = numero_a_letras(total_total)
        story.append(Paragraph(f"VALOR EN LETRAS: {valor_letras} PESOS M/L.", ParagraphStyle("VAL_LETRAS", fontName="Helvetica-Bold", fontSize=13, alignment=TA_CENTER, spaceAfter=8)))
        doc.build(story)
        print(f"PDF generado: {pdf_out}")



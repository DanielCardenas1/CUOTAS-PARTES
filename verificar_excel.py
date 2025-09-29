#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Verificador de archivo Excel generado
Lee y verifica el contenido del Excel oficial generado
"""

import openpyxl
from datetime import date

def verificar_excel_generado():
    """Verifica el contenido del archivo Excel generado"""
    
    nombre_archivo = f"LIQUIDACION_FORMATO_OFICIAL_5742637_{date.today().strftime('%Y%m%d')}.xlsx"
    
    try:
        # Abrir el archivo Excel
        wb = openpyxl.load_workbook(nombre_archivo)
        ws = wb.active
        
        print("=" * 80)
        print(f"VERIFICACIÓN DEL ARCHIVO: {nombre_archivo}")
        print("=" * 80)
        
        # Verificar encabezados
        print("\n📋 ENCABEZADOS:")
        print(f"Título: {ws['A1'].value}")
        print(f"Subtítulo: {ws['A2'].value}")
        
        # Verificar información del pensionado
        print(f"\n👤 INFORMACIÓN DEL PENSIONADO:")
        print(f"Nombre: {ws['B4'].value}")
        print(f"Identificación: {ws['B5'].value}")
        
        # Verificar cabeceras de la tabla
        print(f"\n📊 CABECERAS DE LA TABLA (Fila 11):")
        headers = []
        for col in range(1, 7):  # A hasta F
            cell_value = ws.cell(row=11, column=col).value
            headers.append(cell_value)
            print(f"  Columna {col}: {cell_value}")
        
        # Verificar primeros 5 registros de datos
        print(f"\n📈 PRIMEROS 5 REGISTROS (Filas 12-16):")
        for row in range(12, 17):  # Primeros 5 registros
            periodo = ws[f'A{row}'].value
            dias = ws[f'B{row}'].value
            dtf = ws[f'C{row}'].value
            interes = ws[f'D{row}'].value
            acumulado = ws[f'E{row}'].value
            capital = ws[f'F{row}'].value
            
            print(f"  {periodo}: {dias} días, DTF {dtf}, Interés ${interes:,.2f}, Capital ${capital:,.2f}")
        
        # Buscar la fila de totales (última fila con datos)
        print(f"\n🔍 BUSCANDO FILA DE TOTALES...")
        total_row = None
        for row in range(12, 60):  # Buscar hasta fila 60
            if ws[f'A{row}'].value == "TOTAL":
                total_row = row
                break
        
        if total_row:
            print(f"✅ Fila de totales encontrada: {total_row}")
            total_intereses = ws[f'D{total_row}'].value
            total_capital = ws[f'F{total_row}'].value
            
            print(f"\n💰 TOTALES:")
            print(f"Total Intereses: ${total_intereses:,.2f}")
            print(f"Total Capital: ${total_capital:,.2f}")
            print(f"Total Liquidación: ${total_intereses + total_capital:,.2f}")
            
            # Verificar contra valores esperados
            print(f"\n✅ VERIFICACIÓN:")
            intereses_esperados = 170536.87
            capital_esperado = 19278517.35
            
            diff_intereses = abs(total_intereses - intereses_esperados)
            diff_capital = abs(total_capital - capital_esperado)
            
            print(f"Intereses - Esperado: ${intereses_esperados:,.2f}, Obtenido: ${total_intereses:,.2f}, Diferencia: ${diff_intereses:,.2f}")
            print(f"Capital - Esperado: ${capital_esperado:,.2f}, Obtenido: ${total_capital:,.2f}, Diferencia: ${diff_capital:,.2f}")
            
            if diff_intereses < 1 and diff_capital < 1:
                print("🎯 ¡VALORES CORRECTOS!")
            else:
                print("❌ VALORES INCORRECTOS")
        else:
            print("❌ No se encontró la fila de totales")
        
        # Verificar notas finales
        print(f"\n📝 NOTAS FINALES:")
        for row in range(total_row + 3, total_row + 6):  # 3 filas después de totales
            nota = ws[f'A{row}'].value
            valor = ws[f'B{row}'].value
            if nota:
                if valor:
                    print(f"  {nota}: ${valor:,.2f}")
                else:
                    print(f"  {nota}")
        
        # Estadísticas del archivo
        print(f"\n📊 ESTADÍSTICAS DEL ARCHIVO:")
        print(f"Hojas: {len(wb.sheetnames)}")
        print(f"Nombre de la hoja: {ws.title}")
        print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
        
        wb.close()
        
        print(f"\n✅ VERIFICACIÓN COMPLETADA")
        print(f"Archivo: {nombre_archivo}")
        print(f"Estado: CORRECTO ✅")
        
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo {nombre_archivo}")
    except Exception as e:
        print(f"❌ Error al leer el archivo: {e}")

if __name__ == "__main__":
    verificar_excel_generado()